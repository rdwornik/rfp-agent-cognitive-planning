"""
rfp_excel_agent.py
RFP Excel Agent - Processes Excel files with colored cells (green = questions to answer).

This script:
1. Reads Excel files with multiple tabs
2. Detects green cells (FF00FF00) marking questions to answer
3. Uses intelligent header and column detection
4. Generates answers using RAG pipeline (ChromaDB + LLM)
5. Writes answers back to the detected answer column
6. Saves as a new file (never modifies original)

Usage:
    python rfp_excel_agent.py --input test.xlsx --dry-run
    python rfp_excel_agent.py --input test.xlsx
    python rfp_excel_agent.py --input test.xlsx --solution wms_native --model claude
    python rfp_excel_agent.py --input "RFP_Customer.xlsx" --solution planning --model claude --anonymize --workers 8
"""

import argparse
import os
import sys
import json
import time
import concurrent.futures
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Dict, Optional, Any

# openpyxl for Excel manipulation
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.cell import Cell
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# --- PROJECT ROOT AND IMPORTS ---
PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT / "scripts/core"))

from llm_router import LLMRouter
from anonymization import AnonymizationMiddleware

# --- CONFIGURATION ---
PLATFORM_MATRIX_PATH = PROJECT_ROOT / "config/platform_matrix.json"

# Green color constant (ARGB format)
GREEN_COLOR = 'FF00FF00'
GREEN_COLOR_SHORT = '00FF00'


def get_available_solutions() -> List[str]:
    """Load available solutions from platform_matrix.json."""
    if not PLATFORM_MATRIX_PATH.exists():
        return []
    with open(PLATFORM_MATRIX_PATH, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return list(data.get("solutions", {}).keys())


def count_images_and_charts(workbook) -> Dict[str, int]:
    """
    Count images and charts in all sheets.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        Dict with counts per sheet: {sheet_name: count}
    """
    counts = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        image_count = len(getattr(sheet, '_images', []))
        chart_count = len(getattr(sheet, '_charts', []))
        counts[sheet_name] = image_count + chart_count
    return counts


def is_green_cell(cell: Any) -> bool:
    """
    Check if a cell has green fill color (FF00FF00 ONLY).

    IMPORTANT: Only matches exact green FF00FF00, not other shades.

    Args:
        cell: openpyxl Cell object

    Returns:
        True if cell has exact green fill color FF00FF00
    """
    if not cell or not hasattr(cell, 'fill'):
        return False

    fill = cell.fill
    if fill is None:
        return False

    # Check fgColor (foreground color)
    if hasattr(fill, 'fgColor') and fill.fgColor:
        fg = fill.fgColor
        if hasattr(fg, 'rgb') and fg.rgb:
            rgb = str(fg.rgb).upper()
            # ONLY accept exact match: FF00FF00 or 00FF00
            if len(rgb) == 8:
                return rgb == GREEN_COLOR  # FF00FF00 only
            elif len(rgb) == 6:
                return rgb == GREEN_COLOR_SHORT  # 00FF00 only

    # Check bgColor as fallback
    if hasattr(fill, 'bgColor') and fill.bgColor:
        bg = fill.bgColor
        if hasattr(bg, 'rgb') and bg.rgb:
            rgb = str(bg.rgb).upper()
            if len(rgb) == 8:
                return rgb == GREEN_COLOR  # FF00FF00 only
            elif len(rgb) == 6:
                return rgb == GREEN_COLOR_SHORT  # 00FF00 only

    return False


def find_header_row(sheet) -> Optional[int]:
    """
    Find the header row (first row with multiple text values).

    Args:
        sheet: openpyxl worksheet

    Returns:
        Row number (1-indexed) or None if not found
    """
    for row_idx in range(1, min(20, sheet.max_row + 1)):  # Check first 20 rows
        row_values = []
        for col_idx in range(1, min(50, sheet.max_column + 1)):  # Check first 50 columns
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                row_values.append(str(cell.value).strip())

        # Header row should have at least 3 non-empty text cells
        if len(row_values) >= 3:
            return row_idx

    return 1  # Default to first row


def detect_question_column(sheet, header_row: int) -> Tuple[Optional[int], Optional[str]]:
    """
    Detect the question column using priority-based pattern matching.

    Priority:
    1. "Requirement", "Question", "Description"
    2. "Customer Question", "RFP Question", "Functional Requirement"
    3. First column with substantial text content (avg length > 20 chars)

    Args:
        sheet: openpyxl worksheet
        header_row: Row number of the header

    Returns:
        Tuple of (column_index, column_name) or (None, None)
    """
    # Priority 1 patterns
    priority1 = ["requirement", "question", "description"]
    # Priority 2 patterns
    priority2 = ["customer question", "rfp question", "functional requirement"]

    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if cell.value:
            header_text = str(cell.value).strip()
            header_normalized = header_text.lower().replace("_", " ")
            headers[col_idx] = (header_text, header_normalized)

    # Check Priority 1
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority1:
            if pattern in normalized:
                return col_idx, original

    # Check Priority 2
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority2:
            if pattern in normalized:
                return col_idx, original

    # Fallback: First column with substantial text content
    for col_idx in range(1, sheet.max_column + 1):
        total_len = 0
        count = 0
        for row_idx in range(header_row + 1, min(header_row + 20, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                total_len += len(str(cell.value).strip())
                count += 1

        if count > 0 and (total_len / count) > 20:
            header_cell = sheet.cell(row=header_row, column=col_idx)
            header_name = str(header_cell.value).strip() if header_cell.value else f"Column {col_idx}"
            return col_idx, header_name

    return None, None


def detect_answer_column(sheet, header_row: int, question_col: int) -> Tuple[Optional[int], Optional[str]]:
    """
    Detect the answer column using priority-based pattern matching.

    Priority:
    1. "vendor comment", "vendor answer", "vendor response"
    2. "BY comment", "BY answer", "BY response", "Blue Yonder"
    3. "supplier comment", "supplier answer", "supplier response"
    4. "comment", "answer", "response"
    5. First empty column to the right of question column

    Args:
        sheet: openpyxl worksheet
        header_row: Row number of the header
        question_col: Column index of the question column

    Returns:
        Tuple of (column_index, column_name) or (None, None)
    """
    # Priority patterns
    priority1 = ["vendor comment", "vendor answer", "vendor response"]
    priority2 = ["by comment", "by answer", "by response", "blue yonder"]
    priority3 = ["supplier comment", "supplier answer", "supplier response"]
    priority4 = ["comment", "answer", "response"]

    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=header_row, column=col_idx)
        if cell.value:
            header_text = str(cell.value).strip()
            header_normalized = header_text.lower().replace("_", " ")
            headers[col_idx] = (header_text, header_normalized)

    # Check Priority 1
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority1:
            if pattern in normalized:
                return col_idx, original

    # Check Priority 2
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority2:
            if pattern in normalized:
                return col_idx, original

    # Check Priority 3
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority3:
            if pattern in normalized:
                return col_idx, original

    # Check Priority 4
    for col_idx, (original, normalized) in headers.items():
        for pattern in priority4:
            if pattern in normalized:
                return col_idx, original

    # Fallback: First empty column to the right of question column
    for col_idx in range(question_col + 1, sheet.max_column + 2):
        is_empty = True
        for row_idx in range(header_row + 1, min(header_row + 10, sheet.max_row + 1)):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value and str(cell.value).strip():
                is_empty = False
                break

        if is_empty:
            header_cell = sheet.cell(row=header_row, column=col_idx)
            if header_cell.value:
                return col_idx, str(header_cell.value).strip()
            else:
                return col_idx, f"Column_{get_column_letter(col_idx)}"

    return None, None


def scan_green_cells(workbook) -> List[Dict]:
    """
    Scan all tabs in workbook for green cells (FF00FF00 only).

    SAFETY: Detects ANY green cell (FF00FF00) in a row, then processes that row's question.
            Only the answer column will be modified, all other cells preserved.

    Args:
        workbook: openpyxl Workbook object

    Returns:
        List of dicts with: tab_name, row, question_text, question_col, answer_col
    """
    green_cells = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Find header and columns for this sheet
        header_row = find_header_row(sheet)
        question_col, question_col_name = detect_question_column(sheet, header_row)
        answer_col, answer_col_name = detect_answer_column(sheet, header_row, question_col or 1)

        if not question_col:
            continue

        # Scan for green cells (FF00FF00 ONLY) in each row
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            row_has_green = False
            green_col = None

            # Check all cells in the row for green color
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if is_green_cell(cell):
                    row_has_green = True
                    green_col = col_idx
                    break  # Found green cell in this row

            if row_has_green:
                # Get question text from the question column in this row
                question_cell = sheet.cell(row=row_idx, column=question_col)
                question_text = str(question_cell.value).strip() if question_cell.value else ""

                if question_text:
                    green_cells.append({
                        "tab_name": sheet_name,
                        "row": row_idx,
                        "question_text": question_text,
                        "question_col": question_col,
                        "question_col_name": question_col_name,
                        "answer_col": answer_col,
                        "answer_col_name": answer_col_name,
                        "header_row": header_row,
                        "green_cell_col": green_col  # Track which cell was actually green
                    })

    return green_cells


def process_single_question(
    router: LLMRouter,
    question_text: str,
    model: str,
    middleware: AnonymizationMiddleware
) -> str:
    """
    Process a single question through the RAG pipeline.

    Args:
        router: LLMRouter instance
        question_text: The question to answer
        model: LLM model name
        middleware: AnonymizationMiddleware instance

    Returns:
        Generated answer or error message
    """
    if not question_text or len(question_text.strip()) < 3:
        return ""

    try:
        # Anonymize before LLM call
        clean_question, ctx = middleware.before(question_text)

        # Call LLM
        answer = router.generate_answer(clean_question, model=model)

        # De-anonymize response
        final_answer = middleware.after(answer, ctx)

        return final_answer
    except Exception as e:
        return f"ERROR: {str(e)}"


def dry_run_report(input_path: str, workbook, green_cells: List[Dict]) -> None:
    """
    Print dry run report without processing.

    Args:
        input_path: Path to input file
        workbook: openpyxl Workbook object
        green_cells: List of detected green cells
    """
    print("\n" + "=" * 60)
    print("[DRY RUN] Excel Agent Analysis Report")
    print("=" * 60)
    print(f"[INFO] Input file: {input_path}")
    print(f"[INFO] Total tabs scanned: {len(workbook.sheetnames)}")
    print(f"[INFO] Tabs: {', '.join(workbook.sheetnames)}")
    print()

    # Group by tab
    tabs_summary = {}
    for cell_info in green_cells:
        tab = cell_info["tab_name"]
        if tab not in tabs_summary:
            tabs_summary[tab] = {
                "count": 0,
                "questions": [],
                "question_col": cell_info["question_col_name"],
                "answer_col": cell_info["answer_col_name"],
                "header_row": cell_info["header_row"]
            }
        tabs_summary[tab]["count"] += 1
        tabs_summary[tab]["questions"].append(cell_info["question_text"])

    print(f"[INFO] Green cells found: {len(green_cells)} total")
    print()

    for tab_name, summary in tabs_summary.items():
        print(f"[TAB] {tab_name}")
        print(f"      Header row: {summary['header_row']}")
        print(f"      Question column: {summary['question_col']}")
        print(f"      Answer column: {summary['answer_col']}")
        print(f"      Green cells: {summary['count']}")
        print()

        # Show first 3 questions
        print("      Sample questions:")
        for i, q in enumerate(summary["questions"][:3], 1):
            truncated = q[:80] + "..." if len(q) > 80 else q
            print(f"        {i}. {truncated}")

        if len(summary["questions"]) > 3:
            print(f"        ... and {len(summary['questions']) - 3} more")
        print()

    print("=" * 60)
    print("[DRY RUN] No changes made. Use without --dry-run to process.")
    print("=" * 60)


def process_excel_file(
    input_path: str,
    client: str,
    output_path: Optional[str],
    solution: Optional[str],
    model: str,
    anonymize: bool,
    dry_run: bool,
    workers: int
) -> bool:
    """
    Main processing function for Excel files.

    Args:
        input_path: Path to input Excel file
        client: Client name for output filename
        output_path: Path to output file (or None for auto-generated)
        solution: Solution code for platform-aware context
        model: LLM model name
        anonymize: Enable anonymization
        dry_run: Only analyze, don't process
        workers: Number of parallel workers

    Returns:
        True if successful, False otherwise
    """
    # Check file exists
    if not os.path.exists(input_path):
        print(f"[ERROR] File not found: {input_path}")
        return False

    print(f"[INFO] Scanning Excel file: {input_path}")

    # Load workbook with preservation flags
    total_images = 0  # Initialize for later scope
    try:
        # keep_links=True preserves external links
        # NOTE: keep_vba=True is ONLY for .xlsm files and corrupts .xlsx files
        # openpyxl preserves images by default when cells are not modified
        workbook = load_workbook(input_path, keep_links=True)
        print(f"[INFO] Found {len(workbook.sheetnames)} tabs")

        # Count images/charts before processing
        image_counts_before = count_images_and_charts(workbook)
        total_images = sum(image_counts_before.values())
        if total_images > 0:
            print(f"[INFO] Found {total_images} images/charts to preserve")

    except Exception as e:
        print(f"[ERROR] Could not load Excel file: {e}")
        return False

    # Scan for green cells
    print("[INFO] Scanning for green cells...")
    green_cells = scan_green_cells(workbook)

    if not green_cells:
        print("[WARNING] No green cells found in the workbook.")
        print("[INFO] Mark cells with green fill color (FF00FF00) to indicate questions to answer.")
        return True

    print(f"[INFO] Found {len(green_cells)} green cells across {len(set(c['tab_name'] for c in green_cells))} tabs")

    # Dry run mode - just report
    if dry_run:
        dry_run_report(input_path, workbook, green_cells)
        return True

    # Initialize components
    print("[INFO] Initializing LLM Router...")
    router = LLMRouter(solution=solution)
    middleware = AnonymizationMiddleware(enabled=anonymize)

    print()
    print("=" * 60)
    print(f"[INFO] Model: {model.upper()}")
    print(f"[INFO] Solution: {solution.upper() if solution else 'NONE (generic)'}")
    print(f"[INFO] Anonymization: {'ON' if anonymize else 'OFF'}")
    print(f"[INFO] Workers: {workers}")
    print(f"[INFO] Questions to process: {len(green_cells)}")
    print("=" * 60)
    print()

    # Process questions in parallel
    print("[INFO] Processing questions...")
    start_time = time.time()

    # Prepare tasks
    tasks = [(cell_info["question_text"], cell_info) for cell_info in green_cells]
    results = {}

    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        # Submit all tasks
        future_to_cell = {}
        for question_text, cell_info in tasks:
            future = executor.submit(
                process_single_question,
                router,
                question_text,
                model,
                middleware
            )
            future_to_cell[future] = cell_info

        # Collect results with progress
        completed = 0
        for future in concurrent.futures.as_completed(future_to_cell):
            cell_info = future_to_cell[future]
            answer = future.result()

            # Store result keyed by (tab, row)
            key = (cell_info["tab_name"], cell_info["row"])
            results[key] = {
                "answer": answer,
                "answer_col": cell_info["answer_col"]
            }

            completed += 1
            if completed % 5 == 0 or completed == len(tasks):
                pct = completed * 100 / len(tasks)
                print(f"[INFO] Processing question {completed}/{len(tasks)} ({pct:.1f}%)")

    # Write answers back to workbook
    print("[INFO] Writing answers to workbook...")

    # SAFETY: Count total cells and track modifications
    total_cells = 0
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        total_cells += sheet.max_row * sheet.max_column

    modified_cells = 0
    for (tab_name, row), result in results.items():
        sheet = workbook[tab_name]
        answer_col = result["answer_col"]
        answer = result["answer"]

        if answer_col:
            cell = sheet.cell(row=row, column=answer_col)
            cell.value = answer
            modified_cells += 1

    preserved_cells = total_cells - modified_cells

    # Safety logging
    print(f"[SAFETY] Modified {modified_cells} cells (answer column in green-highlighted rows only)")
    print(f"[SAFETY] Preserved {preserved_cells} cells unchanged")

    # Generate output path if not provided
    if not output_path:
        # Format: {client}_{solution}_{model}_{YYYYMMDD}_{HHMM}.xlsx
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        solution_str = solution if solution else "generic"
        output_filename = f"{client}_{solution_str}_{model}_{timestamp}.xlsx"

        # Save to output_rfp_universal/
        output_dir = Path("output_rfp_universal")
        output_dir.mkdir(exist_ok=True)

        output_path = str(output_dir / output_filename)

    # Save workbook
    try:
        workbook.save(output_path)
        print(f"[SUCCESS] Saved to: {output_path}")

        # Verify images/charts were preserved
        try:
            verification_wb = load_workbook(output_path, keep_links=True)
            image_counts_after = count_images_and_charts(verification_wb)
            total_images_after = sum(image_counts_after.values())
            verification_wb.close()

            if total_images > 0:
                if total_images_after == total_images:
                    print(f"[SUCCESS] Preserved all {total_images} images/charts")
                else:
                    print(f"[WARNING] Image count changed: {total_images} -> {total_images_after}")
                    print(f"[WARNING] Some images may not have been preserved correctly")
        except Exception as e:
            print(f"[WARNING] Could not verify image preservation: {e}")

    except Exception as e:
        print(f"[ERROR] Could not save file: {e}")
        return False

    # Summary
    elapsed = time.time() - start_time
    successful = len([r for r in results.values() if r["answer"] and not r["answer"].startswith("ERROR")])
    errors = len([r for r in results.values() if r["answer"].startswith("ERROR")])

    print()
    print("=" * 60)
    print("[SUCCESS] Processing complete!")
    print(f"[INFO] Output: {output_path}")
    print(f"[INFO] Processed: {successful}/{len(results)} questions")
    if errors > 0:
        print(f"[WARNING] Errors: {errors} questions")
    print(f"[INFO] Time: {elapsed:.1f}s ({elapsed/max(1,len(results)):.2f}s/question)")
    print("=" * 60)

    return True


def parse_args():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="RFP Excel Agent - Process Excel files with colored cells",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s --input test.xlsx --client acme --dry-run
  %(prog)s --input test.xlsx --client acme --solution planning --model gemini
  %(prog)s --input "RFP.xlsx" --client ifm --solution planning --model claude --anonymize

Output: output_rfp_universal/{client}_{solution}_{model}_{YYYYMMDD}_{HHMM}.xlsx
Example: output_rfp_universal/ifm_planning_gemini_20250102_1435.xlsx
        """
    )

    parser.add_argument(
        "-i", "--input",
        type=str,
        required=True,
        help="Input Excel file (required)"
    )

    parser.add_argument(
        "-c", "--client",
        type=str,
        required=True,
        help="Client name for output filename (required)"
    )

    parser.add_argument(
        "-o", "--output",
        type=str,
        default=None,
        help="Output file path (optional, auto-generated if not provided)"
    )

    # Get available solutions dynamically
    available_solutions = get_available_solutions()
    parser.add_argument(
        "-s", "--solution",
        type=str,
        default=None,
        choices=available_solutions if available_solutions else None,
        metavar="CODE",
        help=f"Solution code for platform-aware responses. Available: {', '.join(available_solutions[:5])}..." if available_solutions else "Solution code (see platform_matrix.json)"
    )

    parser.add_argument(
        "-m", "--model",
        type=str,
        default="gemini",
        choices=["gemini", "gemini-flash", "claude", "claude-opus", "gpt5", "o3", "deepseek", "deepseek-r1", "kimi", "llama", "grok", "perplexity", "mistral", "qwen", "glm"],
        help="LLM model to use (default: gemini)"
    )

    parser.add_argument(
        "-a", "--anonymize",
        action="store_true",
        help="Enable anonymization of customer data"
    )

    parser.add_argument(
        "-d", "--dry-run",
        action="store_true",
        help="Analyze file without processing or writing output"
    )

    parser.add_argument(
        "-w", "--workers",
        type=int,
        default=4,
        help="Number of parallel workers (default: 4)"
    )

    return parser.parse_args()


def main():
    """Main entry point."""
    # Check openpyxl availability
    if not OPENPYXL_AVAILABLE:
        print("[ERROR] openpyxl library is required. Install with: pip install openpyxl")
        sys.exit(1)

    args = parse_args()

    print()
    print("=" * 60)
    print("RFP Excel Agent")
    print("=" * 60)
    print(f"[INFO] Input: {args.input}")
    print(f"[INFO] Client: {args.client}")
    print(f"[INFO] Output: {args.output or 'Auto-generated'}")
    print(f"[INFO] Model: {args.model}")
    print(f"[INFO] Solution: {args.solution or 'None (generic)'}")
    print(f"[INFO] Anonymize: {'Yes' if args.anonymize else 'No'}")
    print(f"[INFO] Dry run: {'Yes' if args.dry_run else 'No'}")
    print(f"[INFO] Workers: {args.workers}")
    print("=" * 60)

    success = process_excel_file(
        input_path=args.input,
        client=args.client,
        output_path=args.output,
        solution=args.solution,
        model=args.model,
        anonymize=args.anonymize,
        dry_run=args.dry_run,
        workers=args.workers
    )

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
